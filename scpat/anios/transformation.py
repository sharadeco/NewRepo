from flask import Flask, flash, make_response, current_app

import pandas as pd
import os
from openpyxl import load_workbook
from datetime import date, datetime, timedelta
import json
import numpy as np

from .extensions import ray
from .sqlinjection import *

excel_pathhead = current_app.config['DF_TO_EXCEL']

@ray.remote ( memory = 10*1024*1024 )
def checkDFTab(data, tabname):

    if tabname == 'DEMAND':
        path = os.path.join(excel_pathhead, "Anios_temp_demand_data.csv")
    
    else: #if forecast
        path = os.path.join(excel_pathhead, "Anios_temp_forecast_data.csv")

    if 'Code Produit' in data.columns and 'Désignation' in data.columns and 'Date' in data.columns and 'Qté Fact' in data.columns and 'Division' in data.columns and 'Famille de produit' in data.columns and 'Kg sold' in data.columns:
        data1= data[['Division','Désignation','Code Produit','Date','Qté Fact','Famille de produit', 'Kg sold']].copy()
        data1['Date'] = pd.to_datetime(data1['Date'], format='%Y-%m-%d')        
        data1['Qté Fact'].replace(np.nan,0,inplace=True)
        data1['Kg sold'].replace(np.nan,0,inplace=True)
        
        data1['Division'] = data1['Division'].str.upper()
        data1['Division'] = data1['Division'].str.strip()
       
        data1['Famille de produit'] = data1['Famille de produit'].str.strip()

        df = []
        for val in data1['Code Produit']:
            try: 
                x = format(float(val), '.0f')
            except: 
                x = val.strip()
                x = x.lstrip('0')
                x = x.replace(' ','')

            df.append(str(x))

        data1['Code Produit'] = df

        del df 

        data1.to_csv(path, encoding="utf-8", index=False, header = True)
        print("[INFO ]:" + "All the columns in the "+ tabname +" Sheet are present")
        return True

    #Columns do not match
    else: 
        message = "'" + tabname + "' sheet is missing some columns! \nRequired columns: ['Code Produit' ,'Désignation' ,'Date' ,'Qté Fact' ,'Division' ,'Famille de produit', 'Kg sold'] \nProvided columns: "+ str(list(data.columns))
        print("[ERROR ]:" + message)
        flash(message, "danger")
        return False

@ray.remote( memory = 5*1024*1024)
def checkDivisionTab(div, tabname):
    path = os.path.join(excel_pathhead, "Anios_temp_division_data.csv")

    if 'Division' in div.columns and 'Sales Org / Country' in div.columns and 'Code Produit' in div.columns:
        data1= div[['Division','Sales Org / Country','Code Produit']].copy()
        
        data1['Division'] = data1['Division'].str.upper()
        data1['Division'] = data1['Division'].str.strip()
        
        data1['Code Produit'] = data1['Code Produit'].str.upper()
        data1['Code Produit'] = data1['Code Produit'].str.strip()
        
        data1.to_csv(path, encoding="utf-8", index=False, header = True)
        print("[INFO ]:" + "All the columns in the "+ tabname +" Sheet are present")
        return True        
    
    #else Division tab has missing columns
    else:
        
        message= "'" + tabname + "' sheet is missing some columns! \nRequired columns: ['Division', 'Sales Org / Country', 'Code Produit'] \nProvided columns: "+ str(list(div.columns))
        print("[ERROR ]:" + message)
        flash(message, "danger")    
        return False

def checkAniosData(name):
    file= os.path.join(excel_pathhead, name)
    wb=load_workbook(file, read_only=True, keep_links=False) 
    
    if ('Demand' not in wb.sheetnames) and ('DEMAND' not in wb.sheetnames):
        message = "'Demand' sheet is missing in " + name +" file! \nFile contains the sheets : " + wb.sheetnames
        print("[ERROR ]:" + message)
        flash(message, "danger")
        return False
    
    #if the file doesn't contain the Forecast in the sheet
    elif ('Forecast' not in wb.sheetnames) and ('FORECAST' not in wb.sheetnames):
        message = "'Forecast' sheet is missing in " + name +" file!  \nFile contains the sheets : " + wb.sheetnames
        print("[ERROR ]:" + message)
        flash(message, "danger")
        return False
    
    #if the file doesn't contain the division mapping in the sheet
    elif ('Division Mapping' not in wb.sheetnames):
        message="'Division Mapping' sheet is missing in " + name +" file!  \nFile contains the sheets : " + wb.sheetnames
        print("[ERROR ]:" + message)
        flash(message, "danger")
        return False


    #All three sheets are present 
    else: 
        
        try:
            demand=pd.read_excel(file, engine="openpyxl", na_values=['#VALUE!', '#DIV/0!', '#NAME!', '#N/A!'], sheet_name='Demand')
        except:
            demand=pd.read_excel(file, engine="openpyxl", na_values=['#VALUE!', '#DIV/0!', '#NAME!', '#N/A!'], sheet_name='DEMAND')
        
        
        try:
            forecast= pd.read_excel(file,  engine="openpyxl", na_values=['#VALUE!', '#DIV/0!', '#NAME!', '#N/A!'], sheet_name='Forecast')
        except:
            forecast= pd.read_excel(file,  engine="openpyxl", na_values=['#VALUE!', '#DIV/0!', '#NAME!', '#N/A!'], sheet_name='FORECAST')
        
        div= pd.read_excel(file, engine="openpyxl", na_values=['#VALUE!', '#DIV/0!', '#NAME!', '#N/A!'], sheet_name='Division Mapping')


        fdem = ray.get(checkDFTab.remote(demand, 'DEMAND')) == True
        del demand

        ffcs = ray.get(checkDFTab.remote(forecast, "FORECAST")) == True
        del forecast
        
        fdiv = ray.get(checkDivisionTab.remote(div, 'DIVISION')) == True
        del div
        
    return fdem & ffcs & fdiv    #will return true when data from all 3 sheets has been stored in temp file Anios_temp_data.csv


def checkAniosDemandForecast(name):

    st= time.time()
    file =  os.path.join(excel_pathhead, name)
    wb=load_workbook(file, read_only=True, keep_links=False) 
    if (('Data' not in wb.sheetnames) and ('DATA' not in wb.sheetnames))  :
        message = "'Data' sheet is missing in " + name +" file! \nFile contains the sheets : " + wb.sheetnames
        print("[ERROR ]:" + message)
        flash(message, "danger")
        return False

    #if the file doesn't contain the division mapping in the sheet
    elif 'Divisions correspondance' not in wb.sheetnames and 'matching export Jia Jie' not in wb.sheetnames:
        message = "'Divisions correspondance' or 'Matching export Jia Jie' sheet is missing in " + name +" file! \nFile contains the sheets : " + wb.sheetnames
        print("[ERROR ]:" + message)    
        flash(message, "danger")
        return False


    #Data sheet is present in the File
    else:
        
        try:
            data= pd.read_excel(file,  engine="openpyxl", na_values=['#VALUE!', '#DIV/0!', '#NAME!', '#N/A!'], sheet_name='DATA')
        except:
            data= pd.read_excel(file,  engine="openpyxl", na_values=['#VALUE!', '#DIV/0!', '#NAME!', '#N/A!'], sheet_name='Data')
        
        if 'Data type' not in data.columns: 
            message = "'Data type' is missing in " + name +" file! \nFile contains the sheets : " + wb.sheetnames
            print("[ERROR ]:" + message)
            flash(message, "danger")
            return False

        else:
            demand= data[data['Data type'].isin(['demand','Demand','DEMAND'])] 

            forecast= data[data['Data type'].isin(['FORECAST','Forecast','Consensus Fcst','forecast'])] 

            try:
                div=pd.read_excel(file, engine="openpyxl", na_values=['#VALUE!', '#DIV/0!', '#NAME!', '#N/A!'], sheet_name='Divisions correspondance')

            except:            
                div=pd.read_excel(file, engine="openpyxl", na_values=['#VALUE!', '#DIV/0!', '#NAME!', '#N/A!'], sheet_name='matching export Jia Jie')


            fdem = ray.get(checkDFTab.remote(demand, 'DEMAND')) == True
            del demand

            ffcs = ray.get(checkDFTab.remote(forecast, 'FORECAST')) == True
            del forecast
            
            fdiv = ray.get(checkDivisionTab.remote(div, 'DIVISION')) == True
            del div

        return fdem & ffcs & fdiv    #will return true when data from both the sheets will be stored in the temp files



#checking if files are within the allowed formats
def allowed_file(name):

    filename = name.lower()
    flag = False
    
    # for file - checking if 'Anios data.csv' or 'Anios Demand Forecast.xlsx' is provided.
    if ("anios_data" not in filename) and ("anios_new_model" not in filename) :
        message = "File must be either 'Anios data' or 'Anios Demand Forecast' but recieved "+ name +" !"
        print("[ERROR ]:" + message)
        flash(message, "danger")
        return flag
    
    elif ("anios_data" in filename):
        flag = checkAniosData(name)

    elif ("anios_new_model" in filename):
        flag = checkAniosDemandForecast(name)

    return flag


#computing the demand based on the previous month data 
def computeDemand(dem_data, div_data):    
    
    data1 = pd.merge(dem_data,div_data,on ='Division',how ='inner')
    demand_tabledata = data1[['Code Produit_x','Désignation','Code Produit_y','Qté Fact', 'Division','Famille de produit','Date', 'Kg sold']].copy()
    
    del data1
    del dem_data
    del div_data
    
    demand_tabledata['Key'] = demand_tabledata['Code Produit_x'].astype(str) + "*" + demand_tabledata['Famille de produit'] + "*" + demand_tabledata['Code Produit_y']+ "*" + demand_tabledata['Division']

    #filtering the logic of the dataframe to accomodate only previous month data in demand
    demand_tabledata['Date'] = pd.to_datetime(demand_tabledata['Date'], format='%Y-%m-%d')        
    today = date.today()
    if(today.month == 1):
        startdate = datetime(today.year - 1, 12 , 1 )
        enddate = datetime(today.year - 1, 12 , 31)


    else:
        last_day_of_previous_month = today.replace(day=1) - timedelta(days=1)
        startdate = datetime(today.year, today.month -1  , 1)
        enddate = datetime(today.year, today.month -1  , last_day_of_previous_month.day)

    mask=(demand_tabledata['Date'] >= startdate.strftime('%Y-%m-%d')) & (demand_tabledata['Date'] <= enddate.strftime('%Y-%m-%d'))

    # demand_tabledata = demand_tabledata.loc[mask]
    
    df= demand_tabledata[demand_tabledata['Famille de produit'].isin(['MATERIELS','PRODUITS FINIS'])].copy()
    df.dropna(inplace=True)

    del enddate
    del startdate
    del demand_tabledata
    del mask
    del today
    del last_day_of_previous_month
    
    serialized = []
    json_file = os.path.join(excel_pathhead, 'raw_demand.json')
    r = open(json_file,'w')
    with open(json_file, 'a+') as r:
        for i in range(0, len(df), 1000):
            slc = df.iloc[i : i + 1000]
            serialized.append(convert_demand_to_json.remote(slc, json_file))
        
        ray.get(serialized)
        r.close()

        del df
        del r

        flag = False #default value that will if the demand dataframe is false
        if not (len(serialized) == 0):            
            flag = insert_to_db_from_json(json_file, 'DEMAND')
    
    del serialized

    return flag
    


def computeForecast(dem_data, div_data):
    
    data1 = pd.merge(dem_data,div_data,on ='Division',how ='inner')
    fcst_tabledata = data1[['Code Produit_x','Désignation','Code Produit_y','Qté Fact', 'Division','Kg sold', 'Famille de produit','Date']].copy()
    del data1
    del dem_data
    del div_data

    fcst_tabledata['Key'] = fcst_tabledata['Code Produit_x'].astype(str) + "*" + fcst_tabledata['Famille de produit'] + "*" + fcst_tabledata['Code Produit_y']+ "*" + fcst_tabledata['Division']

    #filtering the logic of the dataframe to accomadate only previous month data in demand
    fcst_tabledata['Date'] = pd.to_datetime(fcst_tabledata['Date'], format='%Y-%m-%d')
    today = date.today()

    startdate = datetime(today.year, today.month  , 1)
    enddate = datetime(today.year + 1, 12 , 31)
    
    mask=(fcst_tabledata['Date'] >= startdate.strftime('%Y-%m-%d')) & (fcst_tabledata['Date'] <= enddate.strftime('%Y-%m-%d'))
    fcst_tabledata = fcst_tabledata.loc[mask]
    df= fcst_tabledata[fcst_tabledata['Famille de produit'].isin(['MATERIELS','PRODUITS FINIS'])].copy()
    df.dropna(inplace=True)
    print("In the compute Fcst function *******************************     now")

    del enddate
    del startdate
    del fcst_tabledata
    del mask
    del today

    serialized = []
    json_file = os.path.join(excel_pathhead, 'raw_fcst.json')
    
    r = open(json_file,'w')
    with open(json_file,'a+') as r:
        
        
        for i in range(0, len(df), 1000):
            slc = df.iloc[i : i + 1000]
            serialized.append(convert_fcst_to_json.remote(slc, True, json_file))
        
        ray.get(serialized)
        r.close()

        del df
        del r

        flag = False #default value that will if the demand dataframe is false
        if not (len(serialized) == 0):            
            flag = insert_to_db_from_json(json_file,'FCST')
    
    del serialized
    del json_file
    
    return flag
   
#Function to unpivot the values from the Calculated Forecast values into the table
def computeFCST(data):
    
    dummy = data['Key'].str.partition("*")[2] 
    data['Code Produit'] = data['Key'].str.partition("*")[0] 
    data['Famille de produit'] = dummy.str.partition("*")[0]

    dummy = dummy.str.partition("*")[2]
    data['Division Mapping#Code Produit'] = dummy.str.partition("*")[0]
    data['Sales Div'] = dummy.str.partition("*")[2]

    del dummy
    
    fcst_tabledata= pd.melt(data, id_vars=['Key', 'Code Produit', 'Famille de produit','Division Mapping#Code Produit','Sales Div'], var_name='Date', value_name='Qté Fact')
    print("Hello data changed")
    del data
    
    fcst_tabledata.dropna(inplace=True)

    #filtering the logic of the dataframe to accomodate only previous month data in demand
    fcst_tabledata['Date'] = pd.to_datetime(fcst_tabledata['Date'], format='%Y-%m-%d')
    today = date.today()

    startdate = datetime(today.year, today.month  , 1)
    enddate = datetime(today.year + 1, 12, 31 )

    mask=(fcst_tabledata['Date'] >= startdate.strftime('%Y-%m-%d')) & (fcst_tabledata['Date'] <= enddate.strftime('%Y-%m-%d'))
    fcst_tabledata = fcst_tabledata.loc[mask]
    df= fcst_tabledata[fcst_tabledata['Famille de produit'].isin(['MATERIELS','PRODUITS FINIS'])].copy()
    #print("Df",df)
    del startdate
    del enddate
    del fcst_tabledata
    del mask
    del today

    serialized = []
    json_file = os.path.join(excel_pathhead, 'raw_calfcst.json')
    r =  open(json_file,'w')
    with open(json_file,'a+') as r:
        
        for i in range(0, len(df), 1000):
            slc = df.iloc[i : i + 1000]
            serialized.append(convert_fcst_to_json.remote(slc, False, json_file))
            
        ray.get(serialized)
        r.close()

        flag = False #default value that will if the demand dataframe is false
        if not (len(serialized) == 0):            
            flag = insert_to_db_from_json(json_file, 'CAL')
    
    del serialized
    del json_file

    return flag

def fetch_summary_records(): #This is the summary of the records of calculated and forecast value
    # forecast_summary = fetch_records('FORECAST_SUMMARY')
    # demand_summary = fetch_records('DEMAND_SUMMARY')
    
    # data_val = pd.merge(forecast_summary, demand_summary, on=['Unique Id'], how='outer' )
    
    # data_val['Actual Demand'].replace(np.nan,0,inplace=True)
    # data_val['Actual Forecast'].replace(np.nan,0,inplace=True)
    # data_val['Demand KG'].replace(np.nan,0,inplace=True)
    # data_val['Forecast KG'].replace(np.nan,0,inplace=True)
    # data_val['Model Forecast'].replace(np.nan,0,inplace=True)
    # data_val['Key_x'].fillna(data_val['Key_y'],inplace=True)
    # data_val['Date_x'].fillna(data_val['Date_y'],inplace=True)
    # data_val['Year_x'].fillna(data_val['Year_y'],inplace=True)
    # data_val['Month_x'].fillna(data_val['Month_y'],inplace=True)
    # data_val['Month_x'] = pd.to_datetime(data_val['Month_x'], format='%m').dt.strftime("%B")
    # data_val['Product Code_x'].fillna(data_val['Product Code_y'],inplace=True)
    # data_val['Product Description_x'].fillna(data_val['Product Description_y'],inplace=True)
    # data_val['Material Type_x'].fillna(data_val['Material Type_y'],inplace=True)
    # data_val['Mother Division_x'].fillna(data_val['Mother Division_y'],inplace=True)
    # data_val['Sales Div_x'].fillna(data_val['Sales Div_y'],inplace=True)

    # data_val.drop(columns=['Key_y','Date_y','Year_y','Month_y','Sales Div_y','Product Code_y','Product Description_y','Material Type_y','Mother Division_y'], axis=1, inplace=True)    

    data_val = fetch_records('SUMMARY')
    data_val['Month'] = pd.to_datetime(data_val['Month'], format='%m').dt.strftime("%B")
    data_val.to_csv(os.path.join(excel_pathhead, "dashboard dump.csv"), encoding="utf-8" )
    val = pd.read_csv(os.path.join(excel_pathhead, "dashboard dump.csv"), encoding="utf-8")

    return data_val
